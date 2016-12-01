#!/usr/bin/python

import sys
import operator
import datetime
from Article import Article
from Article_List import Article_List
import dominate
from dominate.tags import *
import glob2
import load_support
from StockInfo import StockInfo
import pandas as pd
import openpyxl
import xlrd
from openpyxl import load_workbook
import calc_support
import quick_run_support
import ticker_reference

#########################################################################
#########################################################################
# Start printing main program menu
def print_main_menu(master_articles, master_tickers, master_stock_data):

    # Main menu banner
    print '\n************************************************************'
    print '**********************    MAIN MENU    *********************'
    print '************************************************************'

    # Program status information
    # Articles loaded for analysis
    if len(master_articles) > 0:
        articles_loaded = 'Articles loaded? Yes\n{0}'.format(master_articles.description)
    else:
        articles_loaded = 'Articles loaded? No'

    # Ticker list created
    if len(master_tickers) > 0:
        sorted_tickers = sorted(master_tickers.items(), key=operator.itemgetter(1))
        sorted_tickers.reverse()
        tickers_identified = 'Ticker list determined? Yes\n'
        tickers_identified += 'Tickers: {0}'.format(', '.join([item[0] for item in sorted_tickers]))
    else:
        tickers_identified = 'Ticker list determined? No'

    # Stock info obtained (and ready for computation)
    ready_tickers = []
    for ticker in master_stock_data:
        for data_set in master_stock_data[ticker]:
            if data_set.data_type == 'base' and data_set.ticker not in ready_tickers:
                ready_tickers.append(data_set.ticker)
    if len(ready_tickers) > 0:
        info_obtained = 'Market data ready for computation? Yes\n'
        info_obtained += 'Data available for: {0}'.format(', '.join(ready_tickers))
    else:
        info_obtained = 'Market data ready for computation? No'

    # Stock info computed (additional columns added to tableframe)
    computed_tickers = []
    for ticker in master_stock_data:
        for data_set in master_stock_data[ticker]:
            if data_set.data_type == 'computed' and data_set.ticker not in computed_tickers:
                computed_tickers.append(data_set.ticker)
    if len(computed_tickers) > 0:
        info_computed = 'Stock metrics computed? Yes\n'
        info_computed += 'Data computed for: {0}\n'.format(', '.join(computed_tickers))
    else:
        info_computed = 'Stock metrics computed? No\n'

    # Print status
    print '\nSTATUS'
    print articles_loaded
    print tickers_identified
    print info_obtained
    print info_computed

    # Ask user what they want to do with the code
    prompt = '\nWHAT DO YOU WANT TO DO?\n' + \
    '[0] Quick run\n' + \
    '[1] Load articles\n' + \
    '[2] Determine most frequent tickers in loaded articles\n' + \
    '[3] Load tickers from file\n' + \
    '[4] Create new ticker list file\n' + \
    '[5] View current ticker list (and associated articles)\n' + \
    '[6] Manually edit current ticker list\n' + \
    '[7] Get financial data for ticker(s)\n' + \
    '[8] Compute stock metrics\n' + \
    '[9] Load stock metrics from Excel\n' + \
    '[10] Exit program\n\n' + \
    'Enter number: '
    try:
        use = int(float(raw_input(prompt)))
    except KeyboardInterrupt:
        print'\n'
        sys.exit()
    except:
        use = -1

    # Return items to main function
    return use

# End printing main program menu
#########################################################################
#########################################################################

#########################################################################
#########################################################################
# Start Option [0] Quick run
def quick_run(master_articles, master_tickers, master_stock_data, master_ticker_reference):

    # Print intro text
    print '\n------------------------------------------------------------'
    print 'Begin [0] Quick run'
    print '------------------------------------------------------------'

    ######################################################
    # Ask user what type of quick run they want to do (in loop until valid selection)
    valid = False
    while valid == False:
        prompt = '\nWhich quick run do you want to do?\n' + \
        '[1] Search streetinsider live -> determine top tickers -> get financial data\n' + \
        '[2] Enter ticker(s) -> get financial data\n' + \
        '[3] Update stock metrics in Excel file (new end date = today)\n' + \
        '[4] Current test configuration (TWX from 2015-01-01 to today)\n\n' + \
        'Enter number: '
        try:
            choice = int(float(raw_input(prompt)))
            valid = choice in range (1,5)
        except KeyboardInterrupt:
            print'\n'
            sys.exit()
        except:
            pass
        if valid == False:
            print '\nInvalid selection. Please try again.'
    # End selection loop
    ######################################################

    # Create dictionary of automated selections that will be passed to functions
    selections = {}
    selections['Quickrun'] = 0
    selections['Days to Search'] = 0
    selections['Num Top Tickers'] = 0
    selections['Tickers'] = []
    selections['Start'] = datetime.datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
    selections['End'] = datetime.datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

    # If user selects option 1
    if choice == 1:
        quick_run_support.run_1(master_articles, master_tickers, master_stock_data, master_ticker_reference, selections)

    # If user selects option 2
    elif choice == 2:
        quick_run_support.run_2(master_articles, master_tickers, master_stock_data, master_ticker_reference, selections)

    # If user selects option 3
    elif choice == 3:
        quick_run_support.run_3(master_articles, master_tickers, master_stock_data, master_ticker_reference, selections)

    # If user selects option 4
    else:
        quick_run_support.run_4(master_articles, master_tickers, master_stock_data, master_ticker_reference, selections)

    # Print success message
    print '\n>>> Successfully completed selected quick run'

    # Print exit text
    print '\n------------------------------------------------------------'
    print 'End [0] Quick run'
    print '------------------------------------------------------------'

# End Option [0] Quick run
#########################################################################
#########################################################################

#########################################################################
#########################################################################
# Start Option [1] Load articles
def load_articles(master_articles, master_tickers, run_type, selections):

    # Print intro text
    print '\n------------------------------------------------------------'
    print 'Begin [1] Load articles (' + run_type + ')'
    print '------------------------------------------------------------'

    # If need to get user input manually (not a quick run)
    if run_type == 'manual':
        ######################################################
        # Ask user if they want to search site(s) live or load articles from previous run (in loop until valid selection)
        valid = False
        while valid == False:
            prompt = '\nHow do you want to load articles?\n' + \
            '[1] Search site(s) live\n' + \
            '[2] Load from previous run\n\n' + \
            'Enter number: '
            try:
                choice = int(float(raw_input(prompt)))
            except KeyboardInterrupt:
                print'\n'
                sys.exit()
            except:
                choice = -1
            if choice in range(1,3):
                valid = True
            if valid == False:
                print '\nInvalid selection. Please try again.'
        # End selection loop
        ######################################################

        # Make sure user doesn't try to select load from previous run if there are
        # no previous runs available
        if choice == 2 and len(glob2.glob('previous-runs/*.txt')) == 0:
            print '\nNo previous runs available to load from. You will need to search site(s) live.'
            choice = 1

    # Else it's a quick run so user inputs already entered
    else:
        choice = 1

    # If user wants to (or needs to) search live
    if choice == 1:
        if run_type == 'manual':
            load_support.load_live(master_articles, 'manual', {})
        else:
            load_support.load_live(master_articles, 'quick', selections)

    # Else if user wants to load articles from previous run
    else:
        load_support.load_previous_run(master_articles)

    # Need to clear out non-user added / previously file loaded top tickers
    for key, value in master_tickers.items():
        if value != 'user added' and value != 'file loaded':
            del master_tickers[key]

    # Print success message
    print '\n>>> Successfully loaded articles'

    # Print exit text
    print '\n------------------------------------------------------------'
    print 'End [1] Load articles (' + run_type + ')'
    print '------------------------------------------------------------'

# End Option [1] Load articles
#########################################################################
#########################################################################

#########################################################################
#########################################################################
# Start Option [2] Determine most frequent tickers in loaded articles
def determine_top_tickers(master_articles, master_tickers, run_type, selections):

    # Print intro text
    print '\n------------------------------------------------------------'
    print 'Begin [2] Determine most frequent tickers in loaded articles (' + run_type + ')'
    print '------------------------------------------------------------'

    # Error check for case that no articles have been loaded
    if len(master_articles) == 0:
        print '\nNo articles loaded. Please load articles and try again.'

    # Else determine user specified number of top tickers
    else:

        # If need to get user input manually (not a quick run)
        if run_type == 'manual':
            ######################################################
            # Ask how many tickers the user would like to know about (in loop until valid selection)
            valid = False
            while valid == False:
                prompt = '\nHow many top tickers would you like to know about? '
                try:
                    num_top_tickers = int(float(raw_input(prompt)))
                    valid = True
                except KeyboardInterrupt:
                    print'\n'
                    sys.exit()
                except:
                    num_top_tickers = -1
                if valid == False:
                    print '\nInvalid selection. Please try again.'
            # End selection loop
            ######################################################

        # Else it's a quick run so user inputs already entered
        else:
            num_top_tickers = selections['Num Top Tickers']

        # Get user specified number of most frequent tickers
        top_tickers = master_articles.return_top(num_top_tickers)

        # Need to clear out non-user added / previously file loaded top tickers
        for key, value in master_tickers.items():
            if value != 'user added' and value != 'file loaded':
                del master_tickers[key]

        # Add tickers to master ticker list
        for key in top_tickers:
            master_tickers[key] = top_tickers[key]

    # Print success message
    print '\n>>> Successfully determined {0} top tickers'.format(num_top_tickers)

    # Print exit text
    print '\n------------------------------------------------------------'
    print 'End [2] Determine most frequent tickers in loaded articles (' + run_type + ')'
    print '------------------------------------------------------------'

# End Option [2] Determine most frequent tickers in loaded articles
#########################################################################
#########################################################################

#########################################################################
#########################################################################
# Start Option [3] Load tickers from file
def load_tickers(master_tickers, run_type, selections):

    # Print intro text
    print '\n------------------------------------------------------------'
    print 'Begin [3] Load tickers from file (' + run_type + ')'
    print '------------------------------------------------------------'

    # Get list of available ticker list files
    ticker_lists = glob2.glob('ticker-lists/*.txt')

    ##############################################################
    # Loop until the user confirms the list that they want to load
    confirmed = False
    while confirmed == False:
        ######################################################
        # Ask user which file they want to  use to load articles (in loop until valid selection)
        valid = False
        while valid == False:
            prompt = '\nWhich ticker list do you want to load?\n'
            n = 0
            for item in ticker_lists:
                n += 1
                prompt += '[{0}] {1}\n'.format(n, item[13:-4])
            prompt += '\nEnter number: '
            try:
                chosen_num = int(float(raw_input(prompt)))
            except KeyboardInterrupt:
                print'\n'
                sys.exit()
            except:
                chosen_num = -1
            if chosen_num in range(1,n+1):
                valid = True
                chosen_file = ticker_lists[chosen_num-1]
            if valid == False:
                print '\nInvalid selection. Please try again.'
        # End selection loop
        ######################################################

        # Read from selected file show tickers to user to confirm this is the correct list
        ticker_file = open(chosen_file, 'r')
        content = ticker_file.readlines()
        content.pop(0)
        prompt = '\nThe following are the tickers in {0}:\n'.format(chosen_file[13:-4])
        n = 0
        while n < len(content):
            n += 1
            prompt += '{1}\n'.format(n, content[n-1].rstrip())
        prompt += '\nDo you want to load these tickers? (Y or N) '
        choice = raw_input(prompt)

        # If confirmed, move on to loading; if not, repeat selection
        if choice == 'Y' or choice == 'Yes':
            confirmed = True

    # End loop
    ##############################################################

    # Need to clear out non-user added / previously file loaded top tickers
    for key, value in master_tickers.items():
        if value != 'user added' and value != 'file loaded':
            del master_tickers[key]

    # Load the tickers into master_tickers
    for line in content:
        master_tickers[line.rstrip().split(' ')[0]] = 'file loaded'

    # Print success message
    print '\n>>> Successfully loaded tickers from: {0}'.format(chosen_file[13:-4])

    # Print exit text
    print '\n------------------------------------------------------------'
    print 'End [3] Load tickers from file (' + run_type + ')'
    print '------------------------------------------------------------'

# End Option [3] Load tickers from file
#########################################################################
#########################################################################

#########################################################################
#########################################################################
# Start Option [4] Create new ticker list file
def create_ticker_file(master_tickers, run_type, selections):

    # Print intro text
    print '\n------------------------------------------------------------'
    print 'Begin [4] Create new ticker list file (' + run_type + ')'
    print '------------------------------------------------------------'

    ######################################################
    # Ask user if they want to edit an existing file or create a new file (in loop until valid selection)
    valid = False
    while valid == False:
        prompt = '\nHow do you want to create a new ticker list file?\n' + \
        '[1] Use all tickers in current ticker list\n' + \
        '[2] Manually enter ticker(s) to use\n\n' + \
        'Enter number: '
        try:
            choice = int(float(raw_input(prompt)))
        except KeyboardInterrupt:
            print'\n'
            sys.exit()
        except:
            choice = -1
        if choice in range(1,3):
            valid = True
        if valid == False:
            print '\nInvalid selection. Please try again.'
    # End selection loop
    ######################################################

    # If the current ticker list is empty, divert the user to option 2
    if choice == 1 and len(master_tickers) == 0:
        print '\nCurrent ticker list is empty. Please enter tickers manually.'
        choice = 2

    # Get tickers to add if the user selects to input tickers themselves
    if choice == 2:
        prompt = '\nEnter the tickers you would like to add to the file: '
        tickers = raw_input(prompt).split(' ')

    # Ask user to name the list, then create a file with that name
    prompt = '\nWhat do you want to name the ticker list? '
    list_name = raw_input(prompt)
    file_name = 'ticker-lists/' + list_name + '.txt'
    f = open(file_name, 'w')

    # Build the strings to write to the ticker list file
    if choice == 1:
        to_write = list_name
        for key in master_tickers:
            to_write += '\n{0}'.format(key)
    if choice == 2:
        to_write = list_name
        for ticker in tickers:
            to_write += '\n{0}'.format(ticker)

    # Write to the file and then close it
    f.write(to_write)
    f.close()

    # Print success message
    print '\n>>> Successfully saved ticker list to: {0}'.format(file_name)

    # Print exit text
    print '\n------------------------------------------------------------'
    print 'End [4] Create new ticker list file (' + run_type + ')'
    print '------------------------------------------------------------'

# End Option [4] Create new ticker list file
#########################################################################
#########################################################################

#########################################################################
#########################################################################
# Start Option [5] View ticker list (and associated articles)
def view_ticker_list(master_articles, master_tickers):

    # Print intro text
    print '\n------------------------------------------------------------'
    print 'Begin [5] View current ticker list (and associated articles)'
    print '------------------------------------------------------------'

    # Error check for case that no articles have been loaded
    if len(master_tickers) == 0:
        print '\nTicker list is empty.'

    # Else display top ticker information
    else:

        # Create ordered list of top tickers
        sorted_tickers = sorted(master_tickers.items(), key=operator.itemgetter(1))
        sorted_tickers.reverse()

        # Display current ticker list and associated article counts
        print '\nCurrent ticker list:'
        n = 1
        for item in sorted_tickers:
            try:
                int(item[1])
                description = str(item[1]) + ' articles'
            except:
                description = item[1]
            print '[{0}] {1:<6} {2}'.format(n, item[0], description)
            n += 1

        # Ask if user would like to see articles associated with tickers
        prompt = '\nDo you want to see the articles associated with each ticker? (Y or N) '
        choice = raw_input(prompt)
        if choice == 'Y' or choice == 'Yes':
            n = 0
            for item in sorted_tickers:
                n += 1
                # For tickers that were added due to article frequency
                if item[1] != 'user added':
                    print '-----------------------------------------'
                    print '[{0}]/[{1}] All {2} articles for {3}'.format(n, len(master_tickers), item[1], item[0])
                    print '-----------------------------------------'
                    print master_articles.all_for_ticker(item[0])
                else:
                    print '-----------------------------------------'
                    print '[{0}]/[{1}] {2} was {3}. No articles to display.'.format(n, len(master_tickers), item[0], item[1])
                    print '-----------------------------------------\n'

    # Print exit text
    print '\n------------------------------------------------------------'
    print 'End [5] View current ticker list (and associated articles)'
    print '------------------------------------------------------------'

# End Option [5] View ticker list (and associated articles)
#########################################################################
#########################################################################

#########################################################################
#########################################################################
# Start Option [6] Manually edit ticker list
def edit_ticker_list(master_tickers, run_type, selections):

    # Print intro text
    print '\n------------------------------------------------------------'
    print 'Begin [6] Manually edit current ticker list (' + run_type + ')'
    print '------------------------------------------------------------'

    # If need to get user input manually (not a quick run)
    if run_type == 'manual':
        # Allow user to edit list as much as they want
        done_editing = False
        while done_editing == False:

            # Create ordered list of top tickers
            sorted_tickers = sorted(master_tickers.items(), key=operator.itemgetter(1))
            sorted_tickers.reverse()

            # Display current ticker list and associated article counts
            print '\nCurrent ticker list:'
            n = 1
            for item in sorted_tickers:
                print '[{0}] {1:<6} {2}'.format(n, item[0], item[1])
                n += 1

            # Ask user how they want to edit the list
            prompt = '\nHow do you want to edit the list? (Add or Delete) '
            choice = raw_input(prompt)

            # If user wants to delete
            if choice == 'Delete':
                prompt = '\nWhat number item do you want to delete? '
                num_to_delete = int(float(raw_input(prompt)))
                if num_to_delete <= len(sorted_tickers):
                    ticker_to_delete = sorted_tickers[num_to_delete-1][0]
                    del master_tickers[ticker_to_delete]
                else:
                    print '\nInvalid selection.'

            # If user wants to add
            elif choice == 'Add':
                prompt = '\nWhat ticker do you want to add? '
                ticker_to_add = raw_input(prompt)
                master_tickers[ticker_to_add] = 'user added'

            # Error checking
            else:
                print '\nInvalid selection.'

            # Ask if the user has any more edits to make
            prompt = '\nDo you want to make any more edits? (Y or N) '
            choice = raw_input(prompt)
            if choice == 'N' or choice == 'No':
                done_editing = True

    # Else it's a quick run so user inputs already entered
    else:
        for ticker_to_add in selections['Tickers']:
            master_tickers[ticker_to_add] = 'user added'

    # Print success message
    print '\n>>> Successfully edited current ticker list'

    # Print exit text
    print '\n------------------------------------------------------------'
    print 'End [6] Manually edit current ticker list (' + run_type + ')'
    print '------------------------------------------------------------'

# End Option [6] Manually edit ticker list
#########################################################################
#########################################################################

#########################################################################
#########################################################################
# Start Option [7] Get financial data for ticker(s)
def get_financial_data(master_tickers, master_stock_data, master_ticker_reference, run_type, selections):

    # Print intro text
    print '\n------------------------------------------------------------'
    print 'Begin [7] Get financial data for ticker(s) (' + run_type + ')'
    print '------------------------------------------------------------'

    # If need to get user input manually (not a quick run)
    if run_type == 'manual':
        ######################################################
        # Ask user what ticker(s) they want to get info for (in loop until valid selection)
        valid = False
        while valid == False:
            prompt = '\nWhat do you want to do?\n' + \
            '[1] Use all tickers in current ticker list\n' + \
            '[2] Manually enter ticker(s) to use\n\n' + \
            'Enter number: '
            try:
                ticker_choice = int(float(raw_input(prompt)))
            except KeyboardInterrupt:
                print'\n'
                sys.exit()
            except:
                ticker_choice = -1
            if ticker_choice in range(1,3):
                valid = True
            if valid == False:
                print '\nInvalid selection. Please try again.'
        # End selection loop
        ######################################################

        # Error check if ticker list is empty
        if ticker_choice == 1 and len(master_tickers) == 0:
            print '\nTicker list is empty. Please add tickers and try again.'

        else:
            # If user wants to manually enter ticker(s) to use
            if ticker_choice == 2:
                prompt = '\nWhat ticker(s) do you want to get data for? '
                tickers_to_use = raw_input(prompt).split(' ')
                # Add the entered tickers to master_tickers
                for ticker in tickers_to_use:
                    master_tickers[ticker] = 'user added'
            # Else if user wants to use all tickers in current ticker list
            else:
                tickers_to_use = [item for item in master_tickers.keys()]

            # Ask user to specify the start date they would like to use
            prompt = '\nEnter start date you would like to use (YYYY-MM-DD): '
            start_date = raw_input(prompt)

            # Change start_date to correct format
            start_year = int(float(start_date.split('-')[0]))
            start_month = int(float(start_date.split('-')[1]))
            start_day = int(float(start_date.split('-')[2]))
            start = datetime.datetime(start_year, start_month, start_day)

            # Ask user to specify the end date they would like to use
            prompt = '\nEnter end date you would like to use (YYYY-MM-DD or \'today\'): '
            end_date = raw_input(prompt)
            # If the user enters 'today'
            if end_date == 'today' or end_date == 'Today':
                end = datetime.datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
            else:
                # Change end_date to correct format
                end_year = int(float(end_date.split('-')[0]))
                end_month = int(float(end_date.split('-')[1]))
                end_day = int(float(end_date.split('-')[2]))
                end = datetime.datetime(end_year, end_month, end_day)

            # Add StockInfo objects to master_stock_data list
            for ticker in tickers_to_use:
                master_stock_data[ticker].append(StockInfo(ticker, master_ticker_reference, start, end))

    # Else it's a quick run so user inputs already entered
    else:
        tickers_to_use = selections['Tickers']
        start = selections['Start']
        end = selections['End']
        for ticker in tickers_to_use:
            master_stock_data[ticker].append(StockInfo(ticker, master_ticker_reference, start, end))

    # Print success message
    print '\n>>> Successfully retrieved data for: {0}'.format(', '.join(tickers_to_use))

    # Print exit text
    print '\n------------------------------------------------------------'
    print 'End [7] Get financial data for ticker(s) (' + run_type + ')'
    print '------------------------------------------------------------'

# End Option [7] Get financial data for ticker(s)
#########################################################################
#########################################################################

#########################################################################
#########################################################################
# Start Option [8] Compute stock metrics
def compute_stock_metrics(master_stock_data, master_ticker_reference, run_type, selections):

    # Print intro text
    print '\n------------------------------------------------------------'
    print 'Begin [8] Compute stock metrics (' + run_type + ')'
    print '------------------------------------------------------------'

    # Error check for case that no data available
    if len(master_stock_data) == 0:
        print '\nNo market data available. Please get data and try again.'
        return

    else:
        # If need to get user input manually (not a quick run)
        if run_type == 'manual':
            ######################################################
            # Ask user what data they want to analyze (in loop until valid selection)
            valid = False
            while valid == False:
                prompt = '\nWhat data do you want to use to compute stock metrics?\n'
                n = 0
                possible_compute = {}
                for ticker in master_stock_data:
                    data_set_index = -1
                    for data_set in master_stock_data[ticker]:
                        data_set_index += 1
                        if data_set.data_type == 'base':
                            n += 1
                            prompt += '[{0}] {1:<6} '.format(n, data_set.ticker)
                            prompt += 'data from {0} to {1}\n'.format(str(data_set.start_date)[:-9], str(data_set.end_date)[:-9])
                            possible_compute[n] = [ticker, data_set_index, data_set]
                prompt += '\nEnter number(s): '
                choices = raw_input(prompt).split(' ')
                try:
                    choices = [int(float(num)) for num in choices]
                    valid = all(num in range(1,n+1) for num in choices)
                except KeyboardInterrupt:
                    print'\n'
                    sys.exit()
                except:
                    pass
                if valid == False:
                    print '\nInvalid selection. Please try again.'
            # End selection loop
            ######################################################

            # Add (valid) choices to list
            chosen_compute = []
            for num in choices:
                chosen_compute.append(possible_compute[num])

        # Else it's a quick run so user inputs already entered
        else:
            chosen_compute = []
            for ticker in master_stock_data:
                if ticker in selections['Tickers']:
                    data_set_index = -1
                    for data_set in master_stock_data[ticker]:
                        data_set_index += 1
                        if data_set.data_type == 'base' and data_set.start_date == selections['Start'] and data_set.end_date == selections['End']:
                            chosen_compute.append([ticker, data_set_index, data_set])

        # Create Excel file to save stock metrics to
        tickers_string = ''
        for item in chosen_compute:
            tickers_string += item[0] + '_'
        temp_data = chosen_compute[0][2]
        last = str(temp_data.data.index[0])[:-9]
        first = str(temp_data.data.index[len(temp_data.data.index) - 1])[:-9]
        excel_file = 'stock-data/' + tickers_string + last + '_to_' + first + '.xlsx'
        writer = pd.ExcelWriter(excel_file)

        # Create worksheet number variable (will be name of each worksheet)
        current_worksheet = 1

        ##############################################################
        ##############################################################
        # Start loop through analysis for all tickers selected
        for chosen in chosen_compute:

            # Easy references for the 3 components of 'chosen'
            current_ticker = chosen[0]
            current_data_set_index = chosen[1]
            current_data_set = chosen[2]

            # df_ticker = ticker for the current object, df_data = all the actual data
            df_ticker = current_data_set.ticker
            df_data = current_data_set.data

            # Reverse the data so it's in decending order
            df_data = df_data.sort_index(axis=0 ,ascending=False)

            # If the dataframe object for the ticker
            if len(df_data.columns.tolist()) == 6:
                # Make new column containing dates (currently in index)
                df_data['Date'] = df_data.index
                cols = df_data.columns.tolist()
                cols.insert(0, cols.pop(cols.index('Date')))
                df_data = df_data.reindex(columns= cols)

                # Make index be a count up from 0
                indices = range(0, len(df_data.index))
                df_data.index = indices

            # Add column to associate each record with ticker
            df_data.insert(0, 'Ticker', current_ticker)

            # Create list with names of additional columns to be added to the table
            columns_to_add = ['Spread', 'Gain', 'Loss', \
            '10 MA', '10 MA %', '15 MA', '15 MA %', '20 MA', '20 MA %', \
            '50 MA', '50 MA %', '100 MA', '100 MA %', '200 MA', '200 MA %', \
            '10 EMA', '20 EMA', '50 EMA', \
            '52 Wk. High', '52 Wk. Low', \
            '14 Sim. Avg. Gain', '14 Exp. Avg. Gain', '14 Sim. Avg. Loss', '14 Exp. Avg. Loss', \
            '28 Sim. Avg. Gain', '28 Exp. Avg. Gain', '28 Sim. Avg. Loss', '28 Exp. Avg. Loss', \
            '42 Sim. Avg. Gain', '42 Exp. Avg. Gain', '42 Sim. Avg. Loss', '42 Exp. Avg. Loss', \
            '56 Sim. Avg. Gain', '56 Exp. Avg. Gain', '56 Sim. Avg. Loss', '56 Exp. Avg. Loss', \
            '14 Sim. RSI', '14 Exp. RSI', '28 Sim. RSI', '28 Exp. RSI',
            '42 Sim. RSI', '42 Exp. RSI', '56 Sim. RSI', '56 Exp. RSI']

            # Initialize additional column values to be 0.0
            for label in columns_to_add:
                 df_data[label] = 0.0

            # Add values to spread column
            df_data ['Spread'] = calc_support.calc_spread(df_data)

            # Add values to gain and loss columns
            df_data['Gain'] = calc_support.calc_gain(df_data)
            df_data['Loss'] = calc_support.calc_loss(df_data)

            # Create lists with days to be used for the various calculations
            MA_days = [10, 15, 20, 50, 100, 200]
            EMA_days = [10, 20, 50]
            RSI_days = [14, 28, 42, 56]

            ##############################################################
            ##############################################################
            # Start loop 1 to fill columns
            for index, row in df_data.iterrows():

                # Getting 52 week high
                df_data.set_value(index, '52 Wk. High', calc_support.calc_high(df_data, index, 52*5))

                # Getting 52 week low
                df_data.set_value(index, '52 Wk. Low', calc_support.calc_low(df_data, index, 52*5))

                # Getting 10, 15, 50, 100, 200 day moving average
                for days in MA_days:
                    column = str(days) + ' MA'
                    df_data.set_value(index, column, calc_support.calc_mov_avg(df_data, index, days))

                # Geting simple average gains and losses to be used later for RSI calcs
                for days in RSI_days:
                    column_gain = str(days) + ' Sim. Avg. Gain'
                    column_loss = str(days) + ' Sim. Avg. Loss'
                    df_data.set_value(index, column_gain, calc_support.calc_avg_gain_s(df_data, index, days))
                    df_data.set_value(index, column_loss, calc_support.calc_avg_loss_s(df_data, index, days))

            # End loop 1 to fill columns
            ##############################################################
            ##############################################################

            ##############################################################
            ##############################################################
            # Start loop 2 to fill columns
            for index in df_data.index:

                # Getting 10, 15, 50, 100, 200 day percent gains
                for days in MA_days:
                    column = str(days) + ' MA %'
                    df_data.set_value(index, column, calc_support.calc_per_gain(df_data, index, days))

            # End loop 2 to fill columns
            ##############################################################
            ##############################################################

            # Flip table prior to calculating EMAs and exponential avg. gains and losses
            df_data = df_data.sort_index(axis=0 ,ascending=False)

            ##############################################################
            ##############################################################
            # Start loop 3 to fill columns
            for index in df_data.index:

                # Getting 10, 20, and 50 day EMA values
                for days in EMA_days:
                    column = str(days) + ' EMA'
                    df_data.set_value(index, column, calc_support.calc_ema(df_data, index, days))

                # Geting exponential average gains and losses to be used later for RSI calcs
                for days in RSI_days:
                    column_gain = str(days) + ' Exp. Avg. Gain'
                    column_loss = str(days) + ' Exp. Avg. Loss'
                    df_data.set_value(index, column_gain, calc_support.calc_avg_gain_e(df_data, index, days))
                    df_data.set_value(index, column_loss, calc_support.calc_avg_loss_e(df_data, index, days))

            # End loop 3 to fill columns
            ##############################################################
            ##############################################################

            # Flip table back to be ordered newest date to oldest date
            df_data = df_data.sort_index(axis=0 ,ascending=True)

            ##############################################################
            ##############################################################
            # Start loop 4 to fill columns
            for index, row in df_data.iterrows():

                # Getting 14, 28, 42, 56 day simple and exponential RSIs
                for days in RSI_days:
                    column_s = str(days) + ' Sim. RSI'
                    column_e = str(days) + ' Exp. RSI'
                    df_data.set_value(index, column_s, calc_support.calc_rsi_s(df_data, index, days))
                    df_data.set_value(index, column_e, calc_support.calc_rsi_e(df_data, index, days))

            # End loop 4 to fill columns
            ##############################################################
            ##############################################################

            # Save dataframe object to a new worksheet in the Excel file
            if current_worksheet == 1:
                worksheet_name = 'Main'
            else:
                worksheet_name = str(current_worksheet)
            df_data.to_excel(writer, worksheet_name)
            print '\nData for {0} saved to worksheet {1} in {2}'.format(current_ticker, current_worksheet, excel_file)
            current_worksheet += 1

            # Re-assign data in StockInfo object as updated dataframe table
            master_stock_data[current_ticker][current_data_set_index].add_metrics(df_data)

        # End loop through analysis for all tickers selected
        ##############################################################
        ##############################################################

        # Save the output Excel file
        writer.save()

    # Add blank worksheets out to 30 total
    wb = load_workbook(excel_file)
    while current_worksheet <= 30:
        wb.create_sheet(str(current_worksheet))
        current_worksheet += 1

    # Add worksheet at front of Excel that has a list of all the tickers
    ticker_sheet = wb.create_sheet('Tickers')

    current_row = 1
    for chosen in chosen_compute:
        current_ticker = chosen[0]
        _ = ticker_sheet.cell(column=1, row=current_row, value=current_ticker)
        current_row += 1    

    # Re-order worksheets to get ticker sheet at the beginning
    my_order = [30, 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, \
    16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29]
    wb._sheets = [wb._sheets[i] for i in my_order]
    wb.save(excel_file)

    # Print success message
    print '\n>>> Successfully computed stock metrics'

    # Print exit text
    print '\n------------------------------------------------------------'
    print 'End [8] Compute stock metrics (' + run_type + ')'
    print '------------------------------------------------------------'

# End Option [8] Compute stock metrics
#########################################################################
#########################################################################

#########################################################################
#########################################################################
# Start Option [9] Load stock metrics from Excel
def load_stock_metrics(master_stock_data, master_ticker_reference, run_type, selections):

    # Print intro text
    print '\n------------------------------------------------------------'
    print 'Begin [9] Load stock metrics from Excel'
    print '------------------------------------------------------------'

    # Get list of available Excel files from previous computations
    available_metrics = glob2.glob('stock-data/*.xlsx')

    ######################################################
    # Ask user which file they want to use to load stock metrics (in loop until valid selection)
    valid = False
    while valid == False:
        prompt = '\nWhich stock metric Excel file do you want to load from?\n'
        n = 0
        for file_name in available_metrics:
            n += 1
            file_details = file_name.split('_')
            file_details.reverse()
            end_date = file_details[0][:-5]
            start_date = file_details[2]
            file_details[len(file_details)-1] = file_details[len(file_details)-1][11:]
            ticker_list = ', '.join(file_details[3:len(file_details)])
            prompt += '[{0}] Tickers: {1}\n'.format(n, ticker_list)
            if n < 10:
                space = '    '
            else:
                space = '     '
            prompt += '{0}From: {1}\tTo: {2}\n\n'.format(space, start_date, end_date)
        prompt += 'Enter number: '
        try:
            chosen_num = int(float(raw_input(prompt)))
        except KeyboardInterrupt:
            print'\n'
            sys.exit()
        except:
            chosen_num = -1
        if chosen_num in range(1,n+1):
            valid = True
            chosen_file = available_metrics[chosen_num-1]
        if valid == False:
            print '\nInvalid selection. Please try again.'
    # End selection loop
    ######################################################

    # Change the dates in the chosen file name to be datetime objects
    chosen_file_details = chosen_file.split('_')
    chosen_file_details.reverse()
    start_date = chosen_file_details[2]
    start_year = int(float(start_date.split('-')[0]))
    start_month = int(float(start_date.split('-')[1]))
    start_day = int(float(start_date.split('-')[2]))
    start = datetime.datetime(start_year, start_month, start_day)
    end_date = chosen_file_details[0][:-5]
    end_year = int(float(end_date.split('-')[0]))
    end_month = int(float(end_date.split('-')[1]))
    end_day = int(float(end_date.split('-')[2]))
    end = datetime.datetime(end_year, end_month, end_day)

    # Open the chosen Excel file and load data into new StockInfo objects
    # Then add to master_stock_data
    to_read = pd.ExcelFile(chosen_file)
    sheet_names = to_read.sheet_names
    for ticker in sheet_names:
        # Get the current sheet
        current_df = pd.read_excel(open(chosen_file, 'rb'), sheetname=ticker)
        # Create new StockInfo object
        to_add = StockInfo(ticker, master_ticker_reference, start, end)
        to_add.add_metrics(current_df)
        master_stock_data[ticker].append(to_add)

    # Print success message
    print '\n>>> Successfully loaded metrics for: {0}'.format(', '.join(sheet_names))

    # Print exit text
    print '\n------------------------------------------------------------'
    print 'End [9] Load stock metrics from Excel'
    print '------------------------------------------------------------'

# End Option [9] Load stock metrics from Excel
#########################################################################
#########################################################################

#########################################################################
#########################################################################
# Start Option [10] Exit program
def exit_program():

    # Print  text
    print '\n------------------------------------------------------------'
    print 'Adios'
    print '------------------------------------------------------------'

# End Option [10] Exit program
#########################################################################
#########################################################################
